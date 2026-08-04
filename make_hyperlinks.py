"""
엑셀의 '홈페이지' 컬럼(텍스트 URL)을 클릭 가능한 하이퍼링크로 일괄 변환.
- 행 색·다른 서식은 그대로 유지 (해당 셀에 링크 + 파란 밑줄 글꼴만 적용).
- 파일을 제자리에서 수정하므로, 원본 보존이 필요하면 미리 복사해 두세요.
사용법:
    python make_hyperlinks.py                      (기본: 영업대상_점검결과_v2.xlsx)
    python make_hyperlinks.py "C:\\경로\\파일.xlsx"   (특정 파일)
"""
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font

DEFAULT = r"C:\Users\binia\Desktop\영업대상_점검결과_v2.xlsx"
URL_COLUMN_NAME = "홈페이지"
LINK_FONT = Font(color="0563C1", underline="single")


def linkify_ws(ws):
    # 1행(헤더)에서 '홈페이지' 컬럼 위치 찾기
    col = None
    for c in next(ws.iter_rows(min_row=1, max_row=1)):
        if c.value == URL_COLUMN_NAME:
            col = c.column
            break
    if not col:
        return 0
    n = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        v = cell.value
        if isinstance(v, str) and v.strip() and not cell.hyperlink:
            url = v.strip() if "://" in v else "https://" + v.strip()
            cell.hyperlink = url
            cell.font = LINK_FONT
            n += 1
    return n


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT
    wb = load_workbook(path)
    total = sum(linkify_ws(ws) for ws in wb.worksheets)
    wb.save(path)
    print(f"하이퍼링크 변환 완료: {total}개 셀")
    print(f"파일: {path}")


if __name__ == "__main__":
    main()
